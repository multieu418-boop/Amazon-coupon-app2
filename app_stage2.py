import streamlit as st
import pandas as pd
import openpyxl
import re
import math
from io import BytesIO

class CouponExpertEngine:
    @staticmethod
    def get_price_map(df):
        """自动感应价格列和ASIN列"""
        p_keywords = ['price', 'your-price', '价格', 'amount', '零售价']
        a_keywords = ['asin', 'sku-asin', '商品编码', 'product-id']
        t_price, t_asin = None, None
        
        for col in df.columns:
            c_low = str(col).lower()
            if any(k in c_low for k in p_keywords) and t_price is None: t_price = col
            if any(k in c_low for k in a_keywords) and t_asin is None: t_asin = col
        
        if t_asin and t_price:
            df[t_asin] = df[t_asin].astype(str).str.strip().str.upper()
            df[t_price] = pd.to_numeric(df[t_price], errors='coerce').fillna(0.0)
            return df.set_index(t_asin)[t_price].to_dict(), t_price
        return {}, "未找到价格列"

    @staticmethod
    def parse_amazon_errors(error_file, price_map):
        """核心解析：支持读取批注并精准拆分报错块"""
        # 关键：必须 load_workbook(..., data_only=True) 有时会丢失批注，这里不带 data_only
        wb = openpyxl.load_workbook(error_file)
        ws = wb.active
        results = []
        
        # 调试信息：看看一共读了多少行
        # st.sidebar.write(f"调试：Excel共计 {ws.max_row} 行")

        for row in range(10, ws.max_row + 1):
            # 获取原始填写的全量ASIN (第一列)
            raw_asins_str = str(ws.cell(row=row, column=1).value or "")
            if not raw_asins_str or raw_asins_str == "None": continue
            row_asins_all = [a.strip() for a in raw_asins_str.split(';') if a.strip()]

            # 获取第14列 (N列) 的批注
            cell_n = ws.cell(row=row, column=14)
            full_error_text = ""
            if cell_n.comment:
                full_error_text = cell_n.comment.text
            
            # 如果这行没报错，全员标记为 KEEP
            if not full_error_text:
                for sn in row_asins_all:
                    results.append({
                        "row": row, "asin": sn, "type": "KEEP", 
                        "decision": "✅ 正常 (保留)", "original_price": price_map.get(sn, 0.0),
                        "target_p": "N/A", "calc_discount": "原样", "orig_pct": ws.cell(row=row, column=3).value
                    })
                continue

            # 如果有报错，解析报错块
            # 正则提取：匹配所有 ASIN 及其紧随其后的解释文字
            error_blocks = re.findall(r'([A-Z0-9]{10})[\s\n]+(.*?)(?=[A-Z0-9]{10}|$)', full_error_text, re.DOTALL)
            error_dict = {asin.strip(): msg.strip() for asin, msg in error_blocks}

            for sn in row_asins_all:
                current_p = price_map.get(sn, 0.0)
                item = {
                    "row": row, "asin": sn, "original_price": current_p,
                    "target_p": "N/A", "type": "KEEP", "decision": "✅ 正常 (保留)",
                    "calc_discount": "原样", "orig_pct": ws.cell(row=row, column=3).value
                }

                if sn in error_dict:
                    msg = error_dict[sn]
                    if any(x in msg for x in ["没有经验证", "参考价", "历史售价"]):
                        item.update({"type": "REMOVE", "decision": "❌ 无参考价 (剔除)", "calc_discount": "剔除"})
                    elif "要求的净价格" in msg:
                        target_match = re.search(r'(?:价格|price)：?\s*[€\$]?\s*([\d\.]+)', msg)
                        target_p = float(target_match.group(1)) if target_match else 0.0
                        if current_p > 0 and target_p > 0:
                            needed = math.ceil(((current_p - target_p) / current_p) * 100)
                            final_pct = max(5, min(needed, 50))
                            item.update({
                                "type": "ADJUST", "decision": "⚠️ 力度不足 (需处理)", 
                                "target_p": target_p, "calc_discount": f"{final_pct}%",
                                "new_pct_val": final_pct
                            })
                        else:
                            item.update({"type": "ADJUST", "decision": "⚠️ 力度不足 (缺原价)", "calc_discount": "无法计算"})
                results.append(item)
        return results

# --- UI 逻辑 ---
st.set_page_config(page_title="Amazon Coupon 智能决策台", layout="wide")
st.title("⚖️ 阶段 2：报错对齐与决策看板")

if 'decisions' not in st.session_state:
    st.session_state.decisions = {}

with st.sidebar:
    st.header("1. 上传必要文件")
    list_f = st.file_uploader("上传 All Listing Report", type=['txt', 'csv'])
    err_f = st.file_uploader("上传亚马逊报错文件", type=['xlsx'])

if list_f and err_f:
    df_l = pd.read_csv(list_f, sep='\t' if list_f.name.endswith('.txt') else ',')
    p_map, p_col = CouponExpertEngine.get_price_map(df_l)
    
    # 核心解析
    items = CouponExpertEngine.parse_amazon_errors(err_f, p_map)
    
    if not items:
        st.error("❌ 未能在 Excel 中解析到任何 ASIN 数据，请确认 N 列是否有批注（红三角）。")
    else:
        # --- 筛选器 ---
        st.subheader("🔍 快速筛选排查")
        filter_status = st.multiselect(
            "查看状态：", 
            ["❌ 无参考价 (剔除)", "⚠️ 力度不足 (需处理)", "✅ 正常 (保留)"],
            default=["❌ 无参考价 (剔除)", "⚠️ 力度不足 (需处理)"]
        )
        
        filtered = [i for i in items if i['decision'] in filter_status]

        # --- 决策看板 ---
        for i, it in enumerate(filtered):
            with st.container(border=True):
                c1, c2, c3, c4, c5 = st.columns([1, 1, 1, 1, 2])
                c1.write(f"**{it['asin']}**")
                c2.write(f"原价: €{it['original_price']}")
                c3.write(f"要求价: {it['target_p']}")
                c4.info(it['calc_discount'])
                
                # 针对不同类型的决策
                if it['type'] == "REMOVE":
                    c5.warning("系统建议：直接剔除")
                    st.session_state.decisions[it['asin']] = "剔除"
                elif it['type'] == "ADJUST":
                    user_choice = c5.radio(f"决策 (ASIN: {it['asin']})", ["接受修复", "太亏了, 剔除"], key=f"rad_{i}", horizontal=True)
                    st.session_state.decisions[it['asin']] = user_choice
                else:
                    c5.success("正常 ASIN")
                    st.session_state.decisions[it['asin']] = "保留"

        # --- 全量缝合导出 ---
        st.divider()
        if st.button("🚀 生成最终提报序列 (基于我的决策)", type="primary"):
            final_groups = {}
            for it in items:
                asin = it['asin']
                choice = st.session_state.decisions.get(asin, "保留")
                
                if choice in ["保留", "接受修复"]:
                    # 确定最终折扣
                    pct = it.get('new_pct_val', it.get('orig_pct', 5))
                    if pct not in final_groups: final_groups[pct] = []
                    final_groups[pct].append(asin)
            
            for pct, asins in final_groups.items():
                with st.expander(f"📦 新提报组：{pct}% 折扣 (共 {len(asins)} 个)"):
                    st.code(";".join(asins))
